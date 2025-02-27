import sqlite3
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, Reference

# SQLite database file path
db_file = "/Users/sandyaudumala/Netflix/NetflixEngagement.db"

# Get the directory of the database file
db_directory = os.path.dirname(db_file)

# Generate filename with current date
current_date = datetime.now().strftime("%Y-%m-%d")
excel_filename = f"NetflixEngagementWBR_{current_date}.xlsx"
excel_path = os.path.join(db_directory, excel_filename)  # Save in the same directory as DB

# Queries and sheet names
queries = {
    "Avg Watch Time by Plan": """
        SELECT subscription_plan, 
               AVG(daily_watch_time) AS avg_daily_watch_time
        FROM customers
        GROUP BY subscription_plan
        ORDER BY avg_daily_watch_time DESC;
    """,
    "Churn Rate by Plan": """
        SELECT subscription_plan,
               COUNT(CASE WHEN churn_status = 'Yes' THEN 1 END) * 100.0 / COUNT(*) AS churn_rate
        FROM customers
        GROUP BY subscription_plan
        ORDER BY churn_rate DESC;
    """,
    "Engagement Rate by Device": """
        SELECT device_used, 
               AVG(engagement_rate) AS avg_engagement_rate
        FROM customers
        GROUP BY device_used
        ORDER BY avg_engagement_rate DESC;
    """,
    "Satisfaction vs Watch Time": """
        SELECT customer_satisfaction, 
               AVG(daily_watch_time) AS avg_daily_watch_time
        FROM customers
        GROUP BY customer_satisfaction
        ORDER BY customer_satisfaction DESC;
    """
}

# Establish connection to SQLite database
conn = sqlite3.connect(db_file)

# Create an Excel writer
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for sheet_name, query in queries.items():
        df = pd.read_sql_query(query, conn)  # Execute query
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Writing to different sheets

# Close the database connection
conn.close()

# Load the workbook to add charts
wb = openpyxl.load_workbook(excel_path)
summary_sheet = wb.create_sheet("Summary Charts")
chart_positions = [(2, 2), (2, 10), (18, 2), (18, 10), (34, 2), (34, 10), (50, 2), (50, 10)]  # Grid layout
position_index = 0

# Function to add a bar chart
def create_chart(sheet, title, category_col, value_col):
    max_row = sheet.max_row
    
    data = Reference(sheet, min_col=value_col, min_row=1, max_row=max_row)
    categories = Reference(sheet, min_col=category_col, min_row=2, max_row=max_row)
    
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = sheet.cell(row=1, column=value_col).value
    chart.x_axis.title = sheet.cell(row=1, column=category_col).value
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    return chart

# Add charts to respective sheets and collect them for summary
dashboard_charts = []
for sheet_name in wb.sheetnames:
    if sheet_name == "Summary Charts":
        continue
    sheet = wb[sheet_name]
    if sheet.max_row > 1 and sheet.max_column >= 2:
        chart = create_chart(sheet, f"{sheet_name} Overview", category_col=1, value_col=2)
        sheet.add_chart(chart, "E5")  # Add chart in sheet
        dashboard_charts.append(create_chart(sheet, f"{sheet_name} Overview", category_col=1, value_col=2))

# Arrange new chart instances in the summary sheet
for chart in dashboard_charts:
    if position_index < len(chart_positions):
        row, col = chart_positions[position_index]
        summary_sheet.add_chart(chart, f"{chr(64+col)}{row}")  # Position charts in grid layout
        position_index += 1

# Save the Excel file with dashboards
wb.save(excel_path)

print(f"Excel file with dashboards saved at: {excel_path}")